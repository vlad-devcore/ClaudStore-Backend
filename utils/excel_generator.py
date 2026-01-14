# ==========================================
# utils/__init__.py
# ==========================================
# Archivo vacío para que Python reconozca utils como paquete


# ==========================================
# utils/helpers.py
# ==========================================
"""
Funciones auxiliares para el sistema de inventario
"""
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple
from decimal import Decimal
import os
import uuid


def generar_nombre_archivo(extension: str = "jpg") -> str:
    """
    Genera un nombre único para archivos (imágenes)
    
    Args:
        extension (str): Extensión del archivo (sin punto)
    
    Returns:
        str: Nombre único del archivo
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = uuid.uuid4().hex[:8]
    return f"{timestamp}_{unique_id}.{extension}"


def formatear_precio(precio: Decimal) -> str:
    """
    Formatea un precio a formato de moneda
    
    Args:
        precio (Decimal): Precio a formatear
    
    Returns:
        str: Precio formateado (ej: "$1,234.56")
    """
    return f"${precio:,.2f}"


def formatear_fecha(fecha: datetime, formato: str = "%d/%m/%Y %H:%M") -> str:
    """
    Formatea una fecha a string
    
    Args:
        fecha (datetime): Fecha a formatear
        formato (str): Formato de salida
    
    Returns:
        str: Fecha formateada
    """
    return fecha.strftime(formato)


def calcular_porcentaje_ganancia(costo: Decimal, precio_venta: Decimal) -> float:
    """
    Calcula el porcentaje de ganancia
    
    Args:
        costo (Decimal): Costo del producto
        precio_venta (Decimal): Precio de venta
    
    Returns:
        float: Porcentaje de ganancia
    """
    if costo == 0:
        return 0.0
    ganancia = precio_venta - costo
    return float((ganancia / costo) * 100)


def calcular_margen_ganancia(costo: Decimal, precio_venta: Decimal) -> Decimal:
    """
    Calcula el margen de ganancia (diferencia)
    
    Args:
        costo (Decimal): Costo del producto
        precio_venta (Decimal): Precio de venta
    
    Returns:
        Decimal: Margen de ganancia
    """
    return precio_venta - costo


def obtener_primer_dia_mes(fecha: Optional[datetime] = None) -> datetime:
    """
    Obtiene el primer día del mes
    
    Args:
        fecha (datetime): Fecha de referencia (si es None, usa hoy)
    
    Returns:
        datetime: Primer día del mes a las 00:00:00
    """
    if fecha is None:
        fecha = datetime.now()
    return fecha.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def obtener_ultimo_dia_mes(fecha: Optional[datetime] = None) -> datetime:
    """
    Obtiene el último día del mes
    
    Args:
        fecha (datetime): Fecha de referencia (si es None, usa hoy)
    
    Returns:
        datetime: Último día del mes a las 23:59:59
    """
    if fecha is None:
        fecha = datetime.now()
    
    # Obtener el primer día del mes siguiente
    if fecha.month == 12:
        primer_dia_siguiente = fecha.replace(year=fecha.year + 1, month=1, day=1)
    else:
        primer_dia_siguiente = fecha.replace(month=fecha.month + 1, day=1)
    
    # Restar un día para obtener el último día del mes actual
    ultimo_dia = primer_dia_siguiente - timedelta(days=1)
    return ultimo_dia.replace(hour=23, minute=59, second=59, microsecond=999999)


def obtener_rango_periodo_actual() -> Tuple[datetime, datetime]:
    """
    Obtiene el rango de fechas del período actual (mes actual)
    
    Returns:
        Tuple[datetime, datetime]: (fecha_inicio, fecha_fin)
    """
    return obtener_primer_dia_mes(), obtener_ultimo_dia_mes()


def paginar_resultados(resultados: List[Any], page: int = 1, page_size: int = 10) -> Dict[str, Any]:
    """
    Pagina una lista de resultados
    
    Args:
        resultados (List): Lista de resultados
        page (int): Número de página (empieza en 1)
        page_size (int): Tamaño de página
    
    Returns:
        Dict: Diccionario con datos paginados
    """
    total = len(resultados)
    total_pages = (total + page_size - 1) // page_size  # Ceil division
    
    # Validar página
    if page < 1:
        page = 1
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    # Calcular índices
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "data": resultados[start_idx:end_idx]
    }


def validar_imagen(filename: str, extensiones_permitidas: set = {"jpg", "jpeg", "png", "gif", "webp"}) -> bool:
    """
    Valida si un archivo es una imagen válida
    
    Args:
        filename (str): Nombre del archivo
        extensiones_permitidas (set): Set de extensiones permitidas
    
    Returns:
        bool: True si es válida
    """
    if not filename:
        return False
    extension = filename.lower().split('.')[-1]
    return extension in extensiones_permitidas


def guardar_archivo_local(archivo_bytes: bytes, nombre_archivo: str, carpeta: str = "static/images") -> str:
    """
    Guarda un archivo localmente
    
    Args:
        archivo_bytes (bytes): Contenido del archivo
        nombre_archivo (str): Nombre del archivo
        carpeta (str): Carpeta donde guardar
    
    Returns:
        str: Ruta relativa del archivo guardado
    """
    # Crear carpeta si no existe
    os.makedirs(carpeta, exist_ok=True)
    
    # Ruta completa
    ruta_completa = os.path.join(carpeta, nombre_archivo)
    
    # Guardar archivo
    with open(ruta_completa, 'wb') as f:
        f.write(archivo_bytes)
    
    # Retornar ruta relativa para la BD
    return f"/{carpeta}/{nombre_archivo}"


def eliminar_archivo_local(ruta_archivo: str) -> bool:
    """
    Elimina un archivo local
    
    Args:
        ruta_archivo (str): Ruta del archivo
    
    Returns:
        bool: True si se eliminó correctamente
    """
    try:
        # Limpiar la ruta (remover / inicial si existe)
        if ruta_archivo.startswith('/'):
            ruta_archivo = ruta_archivo[1:]
        
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
            return True
        return False
    except Exception as e:
        print(f"Error al eliminar archivo: {e}")
        return False


def convertir_decimal_a_float(obj: Any) -> Any:
    """
    Convierte objetos Decimal a float recursivamente (para JSON)
    
    Args:
        obj: Objeto a convertir
    
    Returns:
        Objeto con Decimals convertidos a float
    """
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: convertir_decimal_a_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convertir_decimal_a_float(item) for item in obj]
    return obj


def cursor_to_dict(cursor) -> List[Dict[str, Any]]:
    """
    Convierte resultados de un cursor a lista de diccionarios
    
    Args:
        cursor: Cursor de psycopg2
    
    Returns:
        List[Dict]: Lista de diccionarios
    """
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def validar_periodo_mes(anio: int, mes: int) -> bool:
    """
    Valida que el año y mes sean válidos
    
    Args:
        anio (int): Año
        mes (int): Mes (1-12)
    
    Returns:
        bool: True si es válido
    """
    if not (2000 <= anio <= 2100):
        return False
    if not (1 <= mes <= 12):
        return False
    return True


def calcular_totales_periodo(ventas: List[Dict], compras: List[Dict], 
                             registros_financieros: List[Dict]) -> Dict[str, Decimal]:
    """
    Calcula totales de un período
    
    Args:
        ventas (List[Dict]): Lista de ventas
        compras (List[Dict]): Lista de compras
        registros_financieros (List[Dict]): Lista de registros financieros
    
    Returns:
        Dict: Diccionario con totales calculados
    """
    total_ventas = sum(Decimal(str(v.get('total', 0))) for v in ventas)
    total_compras = sum(Decimal(str(c.get('total', 0))) for c in compras)
    
    inversion_manual = sum(
        Decimal(str(r.get('monto', 0))) 
        for r in registros_financieros 
        if r.get('tipo') == 'inversion'
    )
    
    gastos_manual = sum(
        Decimal(str(r.get('monto', 0))) 
        for r in registros_financieros 
        if r.get('tipo') == 'gasto'
    )
    
    ganancia_manual = sum(
        Decimal(str(r.get('monto', 0))) 
        for r in registros_financieros 
        if r.get('tipo') == 'ganancia'
    )
    
    ganancia_neta = (total_ventas - total_compras + 
                     ganancia_manual - inversion_manual - gastos_manual)
    
    return {
        'total_ventas': total_ventas,
        'total_compras': total_compras,
        'inversion_manual': inversion_manual,
        'gastos_manual': gastos_manual,
        'ganancia_manual': ganancia_manual,
        'ganancia_neta': ganancia_neta
    }


# ==========================================
# utils/excel_generator.py
# ==========================================
"""
Generador de reportes en Excel para el sistema de inventario
"""
from datetime import datetime
from typing import List, Dict, Any, Optional
from decimal import Decimal
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl no instalado. Instala con: pip install openpyxl")


class ExcelGenerator:
    """Clase para generar reportes en Excel"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
        
        self.wb = Workbook()
        # Eliminar hoja por defecto
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Estilos
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _aplicar_estilo_header(self, ws, row: int, max_col: int):
        """Aplica estilo a la fila de encabezado"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _ajustar_ancho_columnas(self, ws):
        """Ajusta el ancho de las columnas automáticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generar_reporte_periodo(self, periodo_data: Dict[str, Any]) -> bytes:
        """
        Genera un reporte completo del período mensual
        
        Args:
            periodo_data (Dict): Diccionario con datos del período
        
        Returns:
            bytes: Archivo Excel en bytes
        """
        # Hoja 1: Resumen General
        self._crear_hoja_resumen(periodo_data)
        
        # Hoja 2: Top Productos
        if periodo_data.get('top_productos'):
            self._crear_hoja_top_productos(periodo_data['top_productos'])
        
        # Hoja 3: Detalles Financieros
        self._crear_hoja_detalles_financieros(periodo_data)
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    def _crear_hoja_resumen(self, periodo_data: Dict):
        """Crea la hoja de resumen general"""
        ws = self.wb.create_sheet("Resumen General")
        
        # Título
        ws['A1'] = f"Reporte Mensual - {periodo_data.get('mes', '')}/{periodo_data.get('anio', '')}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Información del período
        row = 3
        ws[f'A{row}'] = "Período:"
        ws[f'B{row}'] = f"{periodo_data.get('fecha_inicio', '')} - {periodo_data.get('fecha_fin', '')}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        
        # Sección de Ingresos
        ws[f'A{row}'] = "INGRESOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Ventas totales:"
        ws[f'B{row}'] = float(periodo_data.get('total_ventas', 0))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Ganancias manuales:"
        ws[f'B{row}'] = float(periodo_data.get('ganancia_manual', 0))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        row += 2
        
        # Sección de Egresos
        ws[f'A{row}'] = "EGRESOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Compras totales:"
        ws[f'B{row}'] = float(periodo_data.get('total_compras', 0))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Inversión manual:"
        ws[f'B{row}'] = float(periodo_data.get('total_inversion_manual', 0))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        row += 1
        
        ws[f'A{row}'] = "Gastos manuales:"
        ws[f'B{row}'] = float(periodo_data.get('total_gastos_manual', 0))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        row += 2
        
        # Ganancia Neta
        ws[f'A{row}'] = "GANANCIA NETA:"
        ws[f'B{row}'] = float(periodo_data.get('ganancia_neta', 0))
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row += 2
        
        # Estadísticas
        ws[f'A{row}'] = "ESTADÍSTICAS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Productos vendidos (diferentes):"
        ws[f'B{row}'] = periodo_data.get('cantidad_productos_vendidos', 0)
        row += 1
        
        ws[f'A{row}'] = "Productos nuevos registrados:"
        ws[f'B{row}'] = periodo_data.get('cantidad_productos_registrados', 0)
        
        self._ajustar_ancho_columnas(ws)
    
    def _crear_hoja_top_productos(self, top_productos: List[Dict]):
        """Crea la hoja con el top de productos más vendidos"""
        ws = self.wb.create_sheet("Top Productos")
        
        # Título
        ws['A1'] = "Top 8 Productos Más Vendidos"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Headers
        headers = ['Posición', 'Producto', 'Cantidad Vendida', 'Ingresos Generados']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
        
        self._aplicar_estilo_header(ws, 3, len(headers))
        
        # Datos
        for idx, producto in enumerate(top_productos, start=1):
            row = 3 + idx
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = producto.get('nombre', '')
            ws.cell(row=row, column=3).value = producto.get('cantidad_vendida', 0)
            
            ingresos_cell = ws.cell(row=row, column=4)
            ingresos_cell.value = float(producto.get('ingresos', 0))
            ingresos_cell.number_format = '"$"#,##0.00'
        
        self._ajustar_ancho_columnas(ws)
    
    def _crear_hoja_detalles_financieros(self, periodo_data: Dict):
        """Crea la hoja con detalles financieros"""
        ws = self.wb.create_sheet("Detalles Financieros")
        
        # Título
        ws['A1'] = "Desglose Financiero Detallado"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        # Headers
        headers = ['Concepto', 'Tipo', 'Monto']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
        
        self._aplicar_estilo_header(ws, 3, len(headers))
        
        # Datos
        row = 4
        conceptos = [
            ('Ventas', 'Ingreso', periodo_data.get('total_ventas', 0)),
            ('Compras', 'Egreso', periodo_data.get('total_compras', 0)),
            ('Inversión Manual', 'Egreso', periodo_data.get('total_inversion_manual', 0)),
            ('Gastos Manuales', 'Egreso', periodo_data.get('total_gastos_manual', 0)),
        ]
        
        for concepto, tipo, monto in conceptos:
            ws.cell(row=row, column=1).value = concepto
            ws.cell(row=row, column=2).value = tipo
            monto_cell = ws.cell(row=row, column=3)
            monto_cell.value = float(monto)
            monto_cell.number_format = '"$"#,##0.00'
            row += 1
        
        self._ajustar_ancho_columnas(ws)
    
    def generar_reporte_simple(self, titulo: str, headers: List[str], 
                               datos: List[List[Any]]) -> bytes:
        """
        Genera un reporte simple con headers y datos
        
        Args:
            titulo (str): Título del reporte
            headers (List[str]): Lista de encabezados
            datos (List[List]): Lista de listas con los datos
        
        Returns:
            bytes: Archivo Excel en bytes
        """
        ws = self.wb.create_sheet(titulo)
        
        # Título
        ws['A1'] = titulo
        ws['A1'].font = self.title_font
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        
        # Headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
        
        self._aplicar_estilo_header(ws, 3, len(headers))
        
        # Datos
        for row_idx, row_data in enumerate(datos, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Convertir Decimal a float para Excel
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '"$"#,##0.00'
                elif isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'dd/mm/yyyy hh:mm'
                else:
                    cell.value = value
        
        self._ajustar_ancho_columnas(ws)
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()


# ==========================================
# NOTAS DE USO
# ==========================================
"""
CÓMO USAR UTILS:

1. INSTALAR DEPENDENCIA PARA EXCEL:
   pip install openpyxl

2. USAR HELPERS:
   from utils.helpers import formatear_precio, calcular_porcentaje_ganancia
   
   precio_formateado = formatear_precio(Decimal('1234.56'))  # "$1,234.56"
   porcentaje = calcular_porcentaje_ganancia(100, 150)  # 50.0

3. USAR EXCEL GENERATOR:
   from utils.excel_generator import ExcelGenerator
   
   generator = ExcelGenerator()
   excel_bytes = generator.generar_reporte_periodo(periodo_data)
   
   # Guardar archivo
   with open('reporte.xlsx', 'wb') as f:
       f.write(excel_bytes)
   
   # O enviar como respuesta HTTP
   from fastapi.responses import StreamingResponse
   return StreamingResponse(
       io.BytesIO(excel_bytes),
       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
       headers={"Content-Disposition": "attachment; filename=reporte.xlsx"}
   )

4. PAGINACIÓN:
   from utils.helpers import paginar_resultados
   
   resultados_paginados = paginar_resultados(lista_productos, page=1, page_size=10)
   # Retorna: {total, page, page_size, total_pages, data}

5. GUARDAR IMÁGENES:
   from utils.helpers import generar_nombre_archivo, guardar_archivo_local
   
   nombre = generar_nombre_archivo("jpg")
   ruta = guardar_archivo_local(archivo_bytes, nombre)
"""